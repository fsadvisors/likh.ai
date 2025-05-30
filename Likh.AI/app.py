import os
import streamlit as st
import base64, io, json, pandas as pd
from openai import OpenAI
from PyPDF2 import PdfReader
from PIL import Image, ImageEnhance
from PIL.ImageFilter import UnsharpMask
from openpyxl.styles import Font, PatternFill

# ─── Page config (must be first) ──────
st.set_page_config(layout="wide")

# ─── CSS styling ───────────────────────
st.markdown("""
<style>
[data-testid="stAppViewContainer"] { background: #f4f7fa; }
[data-testid="stSidebar"] { background: #2E3B4E; color: white; }
.stButton>button { background-color: #4CAF50; color: white; border-radius: 5px; }
div[data-testid="stDataFrameContainer"] th { background-color: #D7E4BC !important; }
</style>
""", unsafe_allow_html=True)

# ─── Secure API key loading ────────────
api_key = st.secrets["openai"]["api_key"]
if not api_key.startswith("sk-") or len(api_key) < 40:
    st.error("❌ Invalid OpenAI API key.")
    st.stop()
client = OpenAI(api_key=api_key)

# ─── Sidebar instructions ──────────────
st.sidebar.header("How to use")
st.sidebar.markdown("""
1. **📷 Camera**: Snap a paper invoice.  
2. **⬆️ File/PDF**: Upload an image or text-based PDF.  
3. Click **Process**.  
4. **Edit** fields in the table.  
5. **Download** the Excel sheet.
""", unsafe_allow_html=True)

st.title("🖼️ Likh.AI – Invoice Parser & Editor")

# ─── Configuration ─────────────────────
COLUMNS = [
    "Date", "Invoice No.",
    "Particulars (name of originator)", "Location (of originator)", "GSTIN",
    "Party Name", "Party GSTIN",
    "Item", "MRP", "Qty", "Rate", "Amount", "Total Amount", "Disc Amt.", "IGST Payable", "Grand Total"
]

FUNCTIONS = [
    {
        "name": "extract_invoice",
        "description": "Extract each line item into JSON",
        "parameters": {
            "type": "object",
            "properties": {
                "items": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "date":{"type":"string"},
                            "invoice_no":{"type":"string"},
                            "particulars":{"type":"string"},
                            "location":{"type":"string"},
                            "gstin":{"type":"string"},
                            "party_name":{"type":"string"},
                            "party_gstin":{"type":"string"},
                            "item":{"type":"string"},
                            "mrp":{"type":"number"},
                            "qty":{"type":"number"},
                            "rate":{"type":"number"},
                            "amount":{"type":"number"},
                            "total_amount":{"type":"number"},
                            "disc_amt":{"type":"number"},
                            "igst_payable":{"type":"number"},
                            "grand_total":{"type":"number"}
                        },
                        "required":["date","invoice_no","item","qty"]
                    }
                }
            },
            "required":["items"]
        }
    },
    {
        "name": "extract_headers",
        "description": "Extract invoice‐level header fields",
        "parameters": {
            "type": "object",
            "properties": {
                "date":{"type":"string"},
                "invoice_no":{"type":"string"},
                "originator_name":{"type":"string"},
                "originator_location":{"type":"string"},
                "originator_gstin":{"type":"string"},
                "party_name":{"type":"string"},
                "party_gstin":{"type":"string"},
                "grand_total":{"type":"number"}
            },
            "required":["date","invoice_no"]
        }
    }
]

# ─── Helpers ───────────────────────────
def extract_text_from_pdf(pdf_file):
    rdr = PdfReader(pdf_file)
    return "\n".join(p.extract_text() or "" for p in rdr.pages)

def image_to_base64(img):
    buf = io.BytesIO()
    if img.mode=="RGBA": img=img.convert("RGB")
    img.save(buf,"JPEG")
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()

def preprocess_image(img):
    gray = img.convert("L")
    ctr  = ImageEnhance.Contrast(gray).enhance(2.0)
    shp  = ctr.filter(UnsharpMask(radius=1,percent=150,threshold=1))
    up   = shp.resize((shp.width*2, shp.height*2), Image.BICUBIC)
    return up.convert("RGB")

def call_gpt(fn, content):
    resp = client.chat.completions.create(
        model="gpt-4o",
        messages=[{"role":"user","content":content}],
        functions=FUNCTIONS,
        function_call={"name":fn},
        max_tokens=2048
    )
    fc = resp.choices[0].message.function_call
    return json.loads(fc.arguments) if fc else {}

def call_items_img(img):
    c=[{"type":"text","text":"Extract line items."},
       {"type":"image_url","image_url":{"url":image_to_base64(img)}}]
    return call_gpt("extract_invoice", c).get("items",[])

def call_items_txt(txt):
    return call_gpt("extract_invoice", txt).get("items",[])

def call_hdrs(img=None, txt=None):
    if img:
        c=[{"type":"text","text":"Extract headers."},
           {"type":"image_url","image_url":{"url":image_to_base64(img)}}]
    else:
        c=txt
    return call_gpt("extract_headers", c)

def adjust_headers(h):
    o=h.get("originator_name","") or ""
    p=h.get("party_name","") or ""
    if o and not p:
        h["party_name"]=o
        h["party_gstin"]=h.get("originator_gstin","")
        for k in ["originator_name","originator_location","originator_gstin"]:
            h[k]=""
    return h

def fill_from_headers(df,h):
    mapp={
      "originator_name":"Particulars (name of originator)",
      "originator_location":"Location (of originator)",
      "originator_gstin":"GSTIN",
      "party_name":"Party Name",
      "party_gstin":"Party GSTIN",
      "grand_total":"Grand Total"
    }
    for hk,col in mapp.items():
        if v:=h.get(hk):
            df[col]=df[col].mask(df[col]=="" , v)
    return df

def compute_gt(df):
    amt=pd.to_numeric(df["Amount"],errors="coerce").fillna(0)
    igst=pd.to_numeric(df["IGST Payable"],errors="coerce").fillna(0)
    m=df["Grand Total"].replace("",pd.NA).isna()
    df.loc[m,"Grand Total"]=(amt+igst)[m]
    return df

def build_df(items,hdr):
    hdr=adjust_headers(hdr)
    df=pd.DataFrame(items)
    remap={
      "date":"Date","invoice_no":"Invoice No.",
      "particulars":"Particulars (name of originator)",
      "location":"Location (of originator)","gstin":"GSTIN",
      "party_name":"Party Name","party_gstin":"Party GSTIN",
      "item":"Item","mrp":"MRP","qty":"Qty","rate":"Rate",
      "amount":"Amount","total_amount":"Total Amount",
      "disc_amt":"Disc Amt.","igst_payable":"IGST Payable",
      "grand_total":"Grand Total"
    }
    df=df.rename(columns=remap)
    for c in COLUMNS:
        df[c]=df.get(c,"")
    df=df[COLUMNS]
    for g in ["GSTIN","Party GSTIN"]:
        df[g]=df[g].str.upper().str.extract(r"([0-9A-Z]{15})",expand=False).fillna("")
    df=fill_from_headers(df,hdr)
    df=compute_gt(df)
    return df.applymap(lambda v:v.strip() if isinstance(v,str) else v)

def gen_xlsx(df):
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine='openpyxl') as w:
        df.to_excel(w, sheet_name='Invoice', index=False)
        ws=w.sheets['Invoice']
        hf=PatternFill(fill_type='solid',fgColor='D7E4BC')
        ht=Font(bold=True)
        for idx,_ in enumerate(df.columns, start=1):
            cell=ws.cell(row=1, column=idx)
            cell.fill=hf
            cell.font=ht
            ws.column_dimensions[cell.column_letter].width=max(len(cell.value)+2,12)
    buf.seek(0)
    return buf

# ─── UI ─────────────────────────────────
tab1,tab2=st.tabs(["📷 Camera","⬆️ File/PDF"])
for tab,mode in [(tab1,"cam"),(tab2,"up")]:
    with tab:
        if mode=="cam":
            st.header("Scan via Camera")
            img=st.camera_input("Photo",key=f"{mode}_input")
        else:
            st.header("Upload Invoice")
            img=st.file_uploader("Select file",type=["jpg","png","webp","pdf"],key=f"{mode}_input")

        if st.button("Process",key=f"btn_{mode}"):
            if not img:
                st.warning("Please supply an invoice.")
                continue

            if mode=="up" and img.name.lower().endswith(".pdf"):
                txt=extract_text_from_pdf(img)
                items=call_items_txt(txt)
                hdr=call_hdrs(txt=txt)
            else:
                pil=preprocess_image(Image.open(img))
                items=call_items_img(pil)
                hdr=call_hdrs(img=pil)

            if items:
                st.session_state[f"df_{mode}"]=build_df(items,hdr)

for mode,label in [("cam","Camera"),("up","Upload")]:
    key=f"df_{mode}"
    if key in st.session_state:
        st.subheader(f"Invoice Data ({label})")
        df0=st.session_state[key]
        edited=st.data_editor(df0,key=f"edt_{mode}")
        st.session_state[key]=edited
        xlsx=gen_xlsx(edited)
        st.download_button(
            f"Download Excel ({label})",
            data=xlsx,
            file_name=f"invoice_{mode}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"dload_{mode}"
        )
        st.markdown("---")

st.caption("🔗 Powered by OpenAI GPT-4o Vision | Built with ❤️ using Streamlit by FSA")
